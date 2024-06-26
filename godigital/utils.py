# utils.py
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from calendar import monthrange

def get_day_numbers_for_dayoff(dayoff, year, month):
    day_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
    days_off = []

    if '-' in dayoff:
        start_day, end_day = dayoff.split('-')
        start_day_num = day_map[start_day]
        end_day_num = day_map[end_day]

        current_date = datetime(year, month, 1)
        while current_date.month == month:
            if start_day_num <= current_date.weekday() <= end_day_num:
                days_off.append(current_date.day)
            current_date += timedelta(days=1)
    else:
        day_num = day_map[dayoff]
        current_date = datetime(year, month, 1)
        while current_date.month == month:
            if current_date.weekday() == day_num:
                days_off.append(current_date.day)
            current_date += timedelta(days=1)

    return days_off

def generate_excel(schedules, start_year, start_month):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set the worksheet title
    sheet_title = f"B777 PAX Crew Schedule - {start_month}_{start_year}"  # Replace '/' with '_'
    ws.title = sheet_title

    # Merge cells for the heading
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=26)
    
    # Set the heading content
    heading = f"B777 PAX Crew Schedule - {start_month}/{start_year}"  # Keep '/' for display
    ws['A1'] = heading

    # Apply styles to heading
    header_font = Font(bold=True, size=16)
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Remaining code unchanged...

    # Get the number of days in the current month
    num_days_in_month = monthrange(start_year, start_month)[1]

    # Define the headers with the day name and day number
    start_date = datetime(start_year, start_month, 1)
    day_headers = [
        f"{(start_date.replace(day=day)).strftime('%a')} {day}" 
        for day in range(1, num_days_in_month + 1)
    ]
    headers = ["ID", "Name", "Position"] + day_headers

    ws.append(headers)

    # Apply styles to header row
    header_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Organize schedules by technician
    technician_schedules = {}
    for schedule in schedules:
        tech_id = schedule.technician.emp_id
        if tech_id not in technician_schedules:
            technician_schedules[tech_id] = {
                'name': schedule.technician.name,
                'position': schedule.technician.position,
                'shifts': {day: '' for day in range(1, num_days_in_month + 1)}
            }

        # Determine the days off
        if schedule.dayoff:
            days_off = get_day_numbers_for_dayoff(schedule.dayoff, start_year, start_month)
            for day in days_off:
                technician_schedules[tech_id]['shifts'][day] = 'DayOff'

        # Assign shifts to each day in the month
        for day in range(1, num_days_in_month + 1):
            if not technician_schedules[tech_id]['shifts'][day]:  # Only assign if not already a day off
                technician_schedules[tech_id]['shifts'][day] = schedule.shift

    # Add data to the worksheet
    for tech_id, data in technician_schedules.items():
        row = [
            tech_id,
            data['name'],
            data['position'],
            *[data['shifts'][day] for day in range(1, num_days_in_month + 1)]
        ]
        ws.append(row)

    # Adjust column widths
    # Adjust column widths
    # Adjust column widths
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        col_letter = openpyxl.utils.get_column_letter(min_col)
                        max_length = max(len(str(ws.cell(row=min_row, column=min_col).value)) for row in ws.iter_rows(min_row=min_row, max_row=max_row) for cell in row)
                        ws.column_dimensions[col_letter].width = max_length + 2
            else:
                # For non-merged cells, adjust column width normally
                max_length = max(len(str(cell.value)) for cell in row)
                col_letter = cell.column_letter
                ws.column_dimensions[col_letter].width = max_length + 2

    return wb
